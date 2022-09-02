from openpyxl import load_workbook
import numpy as np
import random
import csv

NUM = 10000
BINS = 0.25

dir_path = "C:/Users/dr2mer05/Desktop/2021RnE/G4data/1GeV/"
load_file_path_mu = dir_path + "mu-1GeV" + ".xlsx"
load_file_path_e = dir_path + "e-1GeV" + ".xlsx"

load_wb_mu = load_workbook(load_file_path_mu, data_only=True)
load_wb_e = load_workbook(load_file_path_e, data_only=True)

load_ws = load_wb_mu["Sheet1"]
load_ws = load_wb_e["Sheet1"]

# r_e을 구해 리스트에 저장하기
Ch1_mu_list = []
get_cells = load_ws["C9":"C"+str(NUM+8)]
for row in get_cells:
    for cell in row:
        Ch1_mu_list.append(cell.value)
Ch2_mu_list = []
get_cells = load_ws["D9":"D"+str(NUM+8)]
for row in get_cells:
    for cell in row:
        Ch2_mu_list.append(cell.value)
re_mu_list = []
for i in range(NUM):
    if Ch2_mu_list[i] == 0:
        continue
    re_mu_list.append(Ch1_mu_list[i] / Ch2_mu_list[i]**2 * 100)

# electron

# r_e을 구해 리스트에 저장하기
Ch1_e_list = []
get_cells = load_ws["C9":"C"+str(NUM+8)]
for row in get_cells:
    for cell in row:
        Ch1_e_list.append(cell.value)
Ch2_e_list = []
get_cells = load_ws["D9":"D"+str(NUM+8)]
for row in get_cells:
    for cell in row:
        Ch2_e_list.append(cell.value)
re_e_list = []
for i in range(NUM):
    if Ch2_e_list[i] == 0:
        continue
    re_e_list.append(Ch1_e_list[i] / Ch2_e_list[i]**2 * 100)

data = []
target = []

for i in range(1000):
    target.append(random.randrange(0, 2))
print(target)

bins = np.arange(0, BINS, 0.001)
for t in target:
    if t == 0:
        sample_list = random.sample(re_mu_list, 1000)
    elif t == 1:
        sample_list = random.sample(re_e_list, 1000)

    hist, bins = np.histogram(sample_list, bins)
    hist = hist.tolist()
    data.append(hist)

with open('G4_1GeV_data.csv', 'w', newline='') as f:
    write = csv.writer(f)
    write.writerows(data)
